"""
Router de exportações — Excel e PDF.
"""

from __future__ import annotations

import io
import logging
from typing import Any

from fastapi import APIRouter, Query, Request, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from backend.sheets.sheets_client import SheetsClient
from backend.sheets.dre_builder import DREBuilder
from backend.sheets.bp_builder import BPBuilder
from backend.sheets.dfc_builder import DFCBuilder

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/export", tags=["export"])


class PDFExportBody(BaseModel):
    """Corpo para exportação PDF."""
    year: int = 2025
    charts: list[str] = []


def _get_sheets_client(request: Request) -> SheetsClient:
    client = getattr(request.app.state, "sheets_client", None)
    if client is None:
        raise HTTPException(
            status_code=503, detail="SheetsClient não inicializado."
        )
    return client


def _df_to_records(df) -> list[dict[str, Any]]:
    if df.empty:
        return []
    return df.fillna("").to_dict(orient="records")


@router.get("/excel")
async def export_excel(request: Request, year: str = Query(default="2025")):
    """
    Gera arquivo Excel formatado com abas DRE, BP e DFC.
    Retorna o arquivo como download.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    sheets_client = _get_sheets_client(request)

    try:
        # Ler dados
        dre_df = DREBuilder(sheets_client).get_dre_data()
        bp_df = BPBuilder(sheets_client).get_bp_data()
        dfc_df = DFCBuilder(sheets_client).get_dfc_data()

        wb = Workbook()

        # ── Estilos ──
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1F4E79", end_color="1F4E79", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        number_format = '#,##0.00'
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        bold_font = Font(name="Calibri", bold=True, size=11)
        normal_font = Font(name="Calibri", size=11)

        def _write_sheet(ws, df, title: str):
            """Escreve um DataFrame numa aba do Excel com formatação."""
            ws.title = title
            if df.empty:
                ws.cell(row=1, column=1, value="Sem dados disponíveis")
                return

            # Cabeçalho
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=str(col_name))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Dados
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                for col_idx, col_name in enumerate(df.columns, 1):
                    val = row[col_name]
                    cell = ws.cell(row=row_idx, column=col_idx)

                    # Tentar converter valor numérico
                    if col_idx > 1:  # Primeira coluna é label
                        try:
                            num_val = float(val)
                            cell.value = num_val
                            cell.number_format = number_format
                        except (ValueError, TypeError):
                            cell.value = str(val) if val != "" else ""
                    else:
                        cell.value = str(val) if val != "" else ""

                    cell.border = thin_border

                    # Negrito para linhas totalizadoras
                    first_col_val = str(row[df.columns[0]])
                    is_bold_line = any(
                        kw in first_col_val
                        for kw in [
                            "Total",
                            "Lucro",
                            "EBITDA",
                            "Receita Líquida",
                            "Resultado",
                            "Variação",
                        ]
                    )
                    cell.font = bold_font if is_bold_line else normal_font

            # Ajustar largura das colunas
            for col_idx, col_name in enumerate(df.columns, 1):
                max_len = max(
                    len(str(col_name)),
                    *(len(str(df.iloc[r, col_idx - 1])) for r in range(len(df)))
                ) if len(df) > 0 else len(str(col_name))
                ws.column_dimensions[
                    ws.cell(row=1, column=col_idx).column_letter
                ].width = min(max_len + 4, 25)

            # Congelar cabeçalho
            ws.freeze_panes = "B2"

        # Escrever abas
        _write_sheet(wb.active, dre_df, "DRE")

        ws_bp = wb.create_sheet()
        _write_sheet(ws_bp, bp_df, "BP")

        ws_dfc = wb.create_sheet()
        _write_sheet(ws_dfc, dfc_df, "DFC")

        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"Relatorio_Financeiro_{year}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            },
        )

    except Exception as exc:
        logger.exception("Erro ao gerar Excel")
        raise HTTPException(status_code=500, detail=str(exc))


# ---------------------------------------------------------------------------
# Template HTML para PDF
# ---------------------------------------------------------------------------
_PDF_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="utf-8">
    <style>
        @page { margin: 2cm; size: A4; }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            color: #1a1a2e;
            font-size: 10pt;
            line-height: 1.4;
        }
        .header {
            background: linear-gradient(135deg, #1F4E79, #2980b9);
            color: white;
            padding: 20px 30px;
            border-radius: 8px;
            margin-bottom: 25px;
        }
        .header h1 { margin: 0; font-size: 22pt; }
        .header p { margin: 5px 0 0; opacity: 0.9; font-size: 11pt; }
        .section-title {
            color: #1F4E79;
            border-bottom: 2px solid #1F4E79;
            padding-bottom: 5px;
            margin-top: 25px;
            font-size: 14pt;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
            font-size: 9pt;
        }
        th {
            background: #1F4E79;
            color: white;
            padding: 8px 6px;
            text-align: center;
            font-weight: 600;
        }
        td {
            padding: 5px 6px;
            border-bottom: 1px solid #e0e0e0;
        }
        tr:nth-child(even) { background: #f8f9fa; }
        .bold-row td { font-weight: 700; background: #eaf2f8 !important; }
        .number { text-align: right; }
        .analysis {
            background: #f0f7ff;
            border-left: 4px solid #1F4E79;
            padding: 15px 20px;
            margin: 20px 0;
            border-radius: 0 8px 8px 0;
        }
        .analysis h3 { color: #1F4E79; margin-top: 0; }
        .footer {
            text-align: center;
            color: #888;
            font-size: 8pt;
            margin-top: 30px;
            border-top: 1px solid #ddd;
            padding-top: 10px;
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>Relatório Financeiro</h1>
        <p>Exercício {{ year }}</p>
    </div>

    {% if analysis %}
    <div class="analysis">
        <h3>Análise</h3>
        <p>{{ analysis }}</p>
    </div>
    {% endif %}

    <h2 class="section-title">DRE — Demonstração do Resultado</h2>
    {{ dre_table }}

    <h2 class="section-title">BP — Balanço Patrimonial</h2>
    {{ bp_table }}

    <h2 class="section-title">DFC — Demonstração de Fluxo de Caixa</h2>
    {{ dfc_table }}

    <div class="footer">
        Gerado automaticamente pelo sistema AF Relatório — {{ timestamp }}
    </div>
</body>
</html>
"""

_BOLD_KEYWORDS = [
    "Total", "Lucro", "EBITDA", "Receita Líquida", "Resultado", "Variação",
    "Ativo Total", "Passivo Total", "Caixa Final",
]


def _df_to_html_table(df) -> str:
    """Converte DataFrame para tabela HTML estilizada."""
    if df.empty:
        return "<p><em>Sem dados disponíveis.</em></p>"

    cols = list(df.columns)
    html = "<table><thead><tr>"
    for col in cols:
        html += f"<th>{col}</th>"
    html += "</tr></thead><tbody>"

    for _, row in df.iterrows():
        label = str(row[cols[0]])
        is_bold = any(kw in label for kw in _BOLD_KEYWORDS)
        row_class = ' class="bold-row"' if is_bold else ""
        html += f"<tr{row_class}>"
        for i, col in enumerate(cols):
            val = row[col]
            if i == 0:
                html += f"<td>{val}</td>"
            else:
                try:
                    num = float(val)
                    formatted = f"{num:,.2f}".replace(",", "X").replace(
                        ".", ","
                    ).replace("X", ".")
                    html += f'<td class="number">{formatted}</td>'
                except (ValueError, TypeError):
                    html += f'<td class="number">{val}</td>'
        html += "</tr>"

    html += "</tbody></table>"
    return html


@router.get("/pdf")
async def export_pdf(
    request: Request,
    year: int = Query(default=2025),
):
    """
    Gera relatório PDF com DRE, BP, DFC e análise.

    Usa Jinja2 para template HTML e weasyprint para conversão a PDF.
    """
    from jinja2 import Template
    from datetime import datetime

    sheets_client = _get_sheets_client(request)

    try:
        # Ler dados
        dre_df = DREBuilder(sheets_client).get_dre_data()
        bp_df = BPBuilder(sheets_client).get_bp_data()
        dfc_df = DFCBuilder(sheets_client).get_dfc_data()

        # Converter DataFrames para tabelas HTML
        dre_table = _df_to_html_table(dre_df)
        bp_table = _df_to_html_table(bp_df)
        dfc_table = _df_to_html_table(dfc_df)

        # Análise IA (stub — pode ser integrado com Anthropic/Gemini)
        analysis = ""

        # Renderizar template
        template = Template(_PDF_TEMPLATE)
        html = template.render(
            year=year,
            analysis=analysis,
            dre_table=dre_table,
            bp_table=bp_table,
            dfc_table=dfc_table,
            timestamp=datetime.now().strftime("%d/%m/%Y %H:%M"),
        )

        # Gerar PDF com weasyprint
        try:
            from weasyprint import HTML

            pdf_buffer = io.BytesIO()
            HTML(string=html).write_pdf(pdf_buffer)
            pdf_buffer.seek(0)
        except ImportError:
            # Fallback: retornar HTML se weasyprint não estiver instalado
            logger.warning(
                "weasyprint não disponível — retornando HTML."
            )
            return StreamingResponse(
                io.BytesIO(html.encode("utf-8")),
                media_type="text/html",
                headers={
                    "Content-Disposition": (
                        f'attachment; filename="Relatorio_{year}.html"'
                    )
                },
            )

        filename = f"Relatorio_Financeiro_{year}.pdf"
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            },
        )

    except Exception as exc:
        logger.exception("Erro ao gerar PDF")
        raise HTTPException(status_code=500, detail=str(exc))
